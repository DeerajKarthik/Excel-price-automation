import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')                       # opens the specific worksheet
sheet = wb['Sheet1']                                             # opens the specific sheet
for row in range(2, sheet.max_row + 1):                          # sheet.max_rows gives us number of rows
    cell = sheet.cell(row, 3)                                    # in format (row, column)
    corrected_price = cell.value * 0.9                           # if price decreased by 10%, price = 90% of original.
    corrected_price_cell = sheet.cell(row, 4)                    # corrected values in new column
    corrected_price_cell.value = corrected_price
    cell2 = sheet.cell(1, 4)
    cell2.value = 'Corrected price'
wb.save('transactions2.xlsx')                                    # creates a new xlsx with corrected values.

